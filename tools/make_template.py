#!/usr/bin/env python3
"""Generate gfy-template.xlsx — the Google Sheet template for the GFY site.

Usage:  python3 tools/make_template.py
Output: tools/gfy-template.xlsx  (upload to Google Drive, open as a Sheet)

Each sheet gets its exact header row (row 1, bold, frozen) plus a few rows of
sample data showing the expected shape. Replace the samples with real data.

Field.deposit, Ledger.settled, Calcutta.collected, and Invites.invited /
Invites.responded are checkbox columns on the site. xlsx has no Google
Sheets checkbox type, so the samples below hold the literal strings
TRUE/FALSE; the README documents the one-time conversion step (select the
column > Insert > Checkbox) after "Save as Google Sheets".

Invites has NO email column (v2.2 P-VAULT: all email addresses live in a
separate, never-published admin vault sheet — outside this template, and
outside this codebase entirely). Invites.status accepts "out" (not
returning, silently suppressed) or "declined" (this season only, shown
under the site's ?admin=1 view, reappears next season) — see F-DECLINED.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

SHEETS = {
    "Info": {
        "headers": ["key", "value"],
        "rows": [
            ["dates", "Aug 14–16"],
            ["course", "Meadow Creek"],
            ["lodging", "Bear Creek Lodge"],
            ["format", "2-day scramble"],
            ["first_tee", "2026-08-15T09:00:00-06:00"],
            ["est_year", "2019"],
            ["calcutta_rake", "10"],
            ["calcutta_basis", "gross"],
            ["deposit_amount", "200"],
            ["payment_handle", "Venmo @gfy-duck"],
        ],
    },
    "Course": {
        "headers": ["hole", "par", "yards"],
        # Real MeadowCreek (New Meadows, ID), white tees — spec §13 C-REAL,
        # checksums par 36/36/72, yards 3094/3007/6101. Green tees BANNED.
        "rows": [
            [1, 4, 319], [2, 5, 469], [3, 4, 407], [4, 3, 124], [5, 4, 357],
            [6, 4, 348], [7, 4, 391], [8, 3, 180], [9, 5, 499],
            [10, 4, 286], [11, 4, 354], [12, 4, 344], [13, 3, 148], [14, 4, 406],
            [15, 5, 433], [16, 4, 352], [17, 3, 151], [18, 5, 533],
        ],
    },
    "Field": {
        # team = captain's name of the player's team; the captain's own row
        # names himself (player == team). since = first GFY year (fill once).
        # strengths (§15) = captain-facing scouting note, shows on the public #draft board
        "headers": ["year", "player", "team", "since", "handicap", "status", "deposit", "paid_date", "strengths"],
        "rows": [
            [2026, "Duck", "Duck", 2019, 8, "In", "TRUE", "", "steady off the tee"],  # captain
            [2026, "Hammer", "Duck", 2019, 10, "In", "TRUE", "", ""],    # partner
            [2026, "Sully", "Sully", 2021, 15, "In", "FALSE", "", ""],   # captain, deposit not in yet
            [2026, "Tank", "Sully", 2026, 20, "In", "TRUE", "", ""],     # partner + rookie (since == this season)
            [2026, "Tex", "Tex", 2019, 18, "In", "TRUE", "", "long drives"],  # captain, third team
            # A next-season payer: rooms go in the order people paid, so
            # paid_date matters more than the row's position on the sheet.
            [2027, "Duck", "", "", "", "", "TRUE", "2026-08-20", ""],
        ],
    },
    "Scores": {
        # One row per TEAM per round, keyed by the captain's name (matches
        # Field.team). Duplicate (team, round) rows merge their hole maps.
        "headers": ["year", "team", "round"] + [f"h{i}" for i in range(1, 19)],
        "rows": [
            [2026, "Duck", 1, 4, 5, 3, 6, 4, 4, 3, 5, 5, 4, 5, 3, 4, 4, 4, 3, 6, 4],
            [2026, "Sully", 1, 5, 4, 4, 5, 4, 5, 3, 4, 5, 4, 6, 3, 5, 4, 4, 4, 5, 5],
            # A round in progress: fill holes as they are played, leave the rest blank.
            [2026, "Tex", 1, 5, 4, 3, 5, 4, 6, 4, 4, 5] + [""] * 9,
        ],
    },
    "Schedule": {
        "headers": ["year", "day", "label", "time", "event", "location"],
        "rows": [
            [2026, "Day One", "Friday", "3:00 pm", "Check in, claim a bed", "Bear Creek Lodge"],
            [2026, "Day One", "Friday", "5:30 pm", "Draw for pairings", "Lodge deck"],
            [2026, "Day Two", "Saturday", "9:00 am", "Round One — shotgun start", "Meadow Creek"],
        ],
    },
    "Pairings": {
        # `start` (P-SHOTGUN, §14.1) is optional — the hole a group starts on
        # for shotgun rounds. Blank is fine (tee-time rounds, like Round Two
        # here) — the site renders exactly as it did before the column
        # existed.
        "headers": ["year", "round", "when", "time", "players", "start"],
        "rows": [
            [2026, "Round One", "Saturday · Shotgun", "9:00 am", "Duck · Hammer", 1],
            [2026, "Round One", "Saturday · Shotgun", "9:00 am", "Sully · Tank", 4],
            [2026, "Round Two", "Sunday · Tee times", "8:30 am", "Leaders out last", ""],
        ],
    },
    "Calcutta": {
        # One row per team lot. No buyback in team mode — owners pay full
        # price to the pot; winnings pay the owner.
        "headers": ["year", "team", "owner", "price", "collected"],
        "rows": [
            [2026, "Duck", "Tex", 120, "TRUE"],
            [2026, "Sully", "Duck", 100, "FALSE"],
            [2026, "Tex", "Sully", 90, "TRUE"],
        ],
    },
    "Payout": {
        "headers": ["year", "place", "share"],
        "rows": [
            [2026, 1, 50],
            [2026, 2, 30],
            [2026, 3, 20],
        ],
    },
    "Ledger": {
        "headers": ["year", "player", "buyin", "won", "settled"],
        "rows": [
            [2026, "Duck", 100, 180, "TRUE"],
            [2026, "Hammer", 100, 40, "FALSE"],
            [2026, "Sully", 100, 0, "FALSE"],
        ],
    },
    "Champions": {
        # place (§15): 1/2/3 — BLANK means 1st (old rows keep working).
        # players: the winning roster, any separator (comma, ·, &, "and").
        "headers": ["year", "champion", "score", "place", "players"],
        "rows": [
            [2025, "Duck", "151 (+7)", "", "Duck · Hammer · Sully · Tank"],
            [2025, "Tex", "153 (+9)", 2, "Tex · Sock · Bear · Crash"],
            [2025, "Moose", "155 (+11)", 3, "Moose · Ghost · Blade · Zeke"],
            [2024, "Hammer", "149 (+5)", "", ""],
            [2019, "Tex", "Inaugural", "", ""],
        ],
    },
    "Shame": {
        "headers": ["year", "award", "player", "detail"],
        "rows": [
            [2026, "Cart incident", "Moose", "Found the one tree on the 7th."],
            [2026, "Most balls lost", "Sully", "Eleven. The creek ate nine of them."],
        ],
    },
    "Invites": {
        # Next-season outreach tracking — separate from who's actually paid
        # (that's Field, above). Invites is the AUTHORITATIVE source for
        # out/declined for the next season WHEN the tab has NEXT-season rows
        # at all — if it's empty/unconfigured for a season, Field's own
        # status column is honored instead (a status value on a Field NEXT
        # row is otherwise ignored, flagged); status=out silently suppresses
        # everyone, status=declined shows under the site's ?admin=1 view and
        # comes back into consideration next season on its own.
        "headers": ["year", "player", "invited", "responded", "status"],
        "rows": [
            [2027, "Sully", "TRUE", "TRUE", ""],     # invited + responded
            [2027, "Tex", "TRUE", "FALSE", ""],      # invited, no reply yet
            [2027, "Bear", "FALSE", "FALSE", "out"], # not returning
        ],
    },
    "Rooms": {
        # Lodging assignments — sheet stores WHO's in WHICH room, nothing
        # else; the site derives the paid-but-unassigned queue, the two-
        # rooms/unpaid/unknown-name health flags, and (?admin=1 only) each
        # player's prior-year room. A name prefixed "guest:" is a non-
        # player lodging guest — exempt from the paid/known-name checks,
        # rendered with the prefix stripped plus a small (guest) mark.
        "headers": ["year", "property", "room", "player"],
        "rows": [
            [2026, "Bear Creek Lodge", "1", "Duck"],
            [2026, "Bear Creek Lodge", "1", "Hammer"],
            [2026, "Bear Creek Lodge", "2", "Sully"],
            [2026, "Bear Creek Lodge", "2", "guest:Pat"],
        ],
    },
}


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; create ours in order
    bold = Font(bold=True)
    for name, spec in SHEETS.items():
        ws = wb.create_sheet(title=name)
        ws.append(spec["headers"])
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"
        for row in spec["rows"]:
            ws.append(row)
    out = Path(__file__).resolve().parent / "gfy-template.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(SHEETS)} sheets)")


if __name__ == "__main__":
    main()
