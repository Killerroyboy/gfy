#!/usr/bin/env python3
"""O-TEMPLATECHECK (spec §19): content-diff both template generators' data
against the committed .xlsx binaries. READ-ONLY — this tool never writes an
xlsx (regeneration is the generators' job). Exit 0 = both in sync; 1 = drift
(printed cell-by-cell); 2 = a file could not be read/imported.

Why not a hash gate: openpyxl stamps timestamps, so the binaries are never
byte-reproducible — git diff on them is noise and md5 can't see real drift.
Run: npm run check-template"""
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.dont_write_bytecode = True
sys.path.insert(0, str(HERE))

def sheet_dict(module_name):
    mod = __import__(module_name)
    for cand in ("SHEETS", "TABS"):
        if hasattr(mod, cand):
            return getattr(mod, cand)
    raise AttributeError(f"{module_name} exposes neither SHEETS nor TABS")

def diff_workbook(xlsx_path, spec, label):
    from openpyxl import load_workbook
    problems = []
    wb = load_workbook(xlsx_path)
    want_sheets = list(spec.keys())
    if wb.sheetnames != want_sheets:
        problems.append(f"{label}: sheet list {wb.sheetnames} != generator {want_sheets}")
    for name, tab in spec.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        want = [tab["headers"]] + [list(r) for r in tab["rows"]]
        got = [[c.value for c in row] for row in ws.iter_rows()]
        norm = lambda v: "" if v is None else str(v)
        for i in range(max(len(want), len(got))):
            w = [norm(v) for v in (want[i] if i < len(want) else [])]
            g = [norm(v) for v in (got[i] if i < len(got) else [])]
            # trailing blanks are xlsx-representation noise, not drift
            while w and w[-1] == "": w.pop()
            while g and g[-1] == "": g.pop()
            if w != g:
                problems.append(f"{label} [{name}] row {i+1}: generator={w} xlsx={g}")
    return problems

def main():
    checks = [
        ("make_template", HERE / "gfy-template.xlsx", "public template"),
        ("make_admin_template", HERE / "gfy-admin-template.xlsx", "admin template"),
    ]
    all_problems = []
    for module, xlsx, label in checks:
        try:
            spec = sheet_dict(module)
        except Exception as e:
            print(f"CANNOT IMPORT {module}: {e}")
            return 2
        try:
            all_problems += diff_workbook(xlsx, spec, label)
        except Exception as e:
            print(f"CANNOT READ {xlsx.name}: {e}")
            return 2
    if all_problems:
        print(f"TEMPLATE DRIFT — {len(all_problems)} difference(s):")
        for p in all_problems:
            print("  " + p)
        print("Fix: edit the GENERATOR, re-run it, commit both. Never hand-edit the xlsx.")
        return 1
    print("templates in sync — both xlsx match their generators cell-for-cell")
    return 0

if __name__ == "__main__":
    sys.exit(main())
